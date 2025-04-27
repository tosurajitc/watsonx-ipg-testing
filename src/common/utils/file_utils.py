"""
Utility functions for file operations in the Watsonx IPG Testing project.

This module provides functions for common file operations such as reading, writing,
parsing various file formats, and handling file paths.
"""

import os
import json
import yaml
import csv
import shutil
import logging
import tempfile
from pathlib import Path
from typing import Dict, List, Any, Union, Optional, BinaryIO, TextIO, Tuple, Set
import xml.etree.ElementTree as ET
from datetime import datetime

# For handling Excel files
try:
    import openpyxl
    import pandas as pd
    EXCEL_SUPPORT = True
except ImportError:
    EXCEL_SUPPORT = False

# Set up logging
logger = logging.getLogger(__name__)


# Path Operations

def ensure_directory_exists(directory_path: str) -> str:
    """
    Ensures that the specified directory exists, creating it if necessary.
    
    Args:
        directory_path: Path to the directory
        
    Returns:
        The absolute path to the directory
        
    Raises:
        OSError: If the directory cannot be created
    """
    abs_path = os.path.abspath(directory_path)
    if not os.path.exists(abs_path):
        try:
            os.makedirs(abs_path, exist_ok=True)
            logger.info(f"Created directory: {abs_path}")
        except OSError as e:
            logger.error(f"Failed to create directory {abs_path}: {e}")
            raise
    return abs_path


def get_file_extension(file_path: str) -> str:
    """
    Get the extension of a file.
    
    Args:
        file_path: Path to the file
        
    Returns:
        The file extension (lowercase, without the dot)
    """
    return os.path.splitext(file_path)[1].lower().lstrip('.')


def list_files(directory_path: str, 
               extensions: Optional[List[str]] = None, 
               recursive: bool = False) -> List[str]:
    """
    List files in a directory, optionally filtering by extension and including subdirectories.
    
    Args:
        directory_path: Path to the directory
        extensions: List of extensions to filter by (without the dot)
        recursive: Whether to search recursively in subdirectories
        
    Returns:
        List of file paths
        
    Raises:
        FileNotFoundError: If the directory does not exist
    """
    if not os.path.exists(directory_path):
        logger.error(f"Directory not found: {directory_path}")
        raise FileNotFoundError(f"Directory not found: {directory_path}")
    
    result = []
    
    if recursive:
        for root, _, files in os.walk(directory_path):
            for file in files:
                file_path = os.path.join(root, file)
                if extensions is None or get_file_extension(file_path) in extensions:
                    result.append(file_path)
    else:
        for item in os.listdir(directory_path):
            file_path = os.path.join(directory_path, item)
            if os.path.isfile(file_path) and (extensions is None or get_file_extension(file_path) in extensions):
                result.append(file_path)
    
    return result


def get_file_info(file_path: str) -> Dict[str, Any]:
    """
    Get information about a file.
    
    Args:
        file_path: Path to the file
        
    Returns:
        Dictionary containing file information
        
    Raises:
        FileNotFoundError: If the file does not exist
    """
    if not os.path.exists(file_path):
        logger.error(f"File not found: {file_path}")
        raise FileNotFoundError(f"File not found: {file_path}")
    
    stat_info = os.stat(file_path)
    
    return {
        'path': os.path.abspath(file_path),
        'name': os.path.basename(file_path),
        'extension': get_file_extension(file_path),
        'size': stat_info.st_size,
        'created': datetime.fromtimestamp(stat_info.st_ctime),
        'modified': datetime.fromtimestamp(stat_info.st_mtime),
        'accessed': datetime.fromtimestamp(stat_info.st_atime),
    }


def create_unique_filename(directory_path: str, 
                           base_name: str, 
                           extension: str) -> str:
    """
    Create a unique filename in the specified directory.
    
    Args:
        directory_path: Path to the directory
        base_name: Base name for the file
        extension: File extension (without the dot)
        
    Returns:
        A unique file path
    """
    ensure_directory_exists(directory_path)
    
    # Clean the base name to remove invalid characters
    base_name = ''.join(c for c in base_name if c.isalnum() or c in ' _-')
    base_name = base_name.strip()
    
    if not base_name:
        base_name = "file"
    
    extension = extension.lstrip('.')
    counter = 0
    
    while True:
        if counter == 0:
            filename = f"{base_name}.{extension}"
        else:
            filename = f"{base_name}_{counter}.{extension}"
        
        file_path = os.path.join(directory_path, filename)
        
        if not os.path.exists(file_path):
            return file_path
        
        counter += 1


# File Reading Operations

def read_file(file_path: str, 
              encoding: str = 'utf-8', 
              binary: bool = False) -> Union[str, bytes]:
    """
    Read a file and return its contents.
    
    Args:
        file_path: Path to the file
        encoding: File encoding (for text files)
        binary: Whether to read the file in binary mode
        
    Returns:
        File contents as string or bytes
        
    Raises:
        FileNotFoundError: If the file does not exist
        IOError: If there's an error reading the file
    """
    if not os.path.exists(file_path):
        logger.error(f"File not found: {file_path}")
        raise FileNotFoundError(f"File not found: {file_path}")
    
    try:
        mode = 'rb' if binary else 'r'
        with open(file_path, mode=mode, encoding=None if binary else encoding) as file:
            return file.read()
    except IOError as e:
        logger.error(f"Error reading file {file_path}: {e}")
        raise


def read_json_file(file_path: str, encoding: str = 'utf-8') -> Dict[str, Any]:
    """
    Read a JSON file and parse its contents.
    
    Args:
        file_path: Path to the JSON file
        encoding: File encoding
        
    Returns:
        Parsed JSON data
        
    Raises:
        FileNotFoundError: If the file does not exist
        json.JSONDecodeError: If the file contains invalid JSON
    """
    content = read_file(file_path, encoding)
    try:
        return json.loads(content)
    except json.JSONDecodeError as e:
        logger.error(f"Error parsing JSON from {file_path}: {e}")
        raise


def read_yaml_file(file_path: str, encoding: str = 'utf-8') -> Dict[str, Any]:
    """
    Read a YAML file and parse its contents.
    
    Args:
        file_path: Path to the YAML file
        encoding: File encoding
        
    Returns:
        Parsed YAML data
        
    Raises:
        FileNotFoundError: If the file does not exist
        yaml.YAMLError: If the file contains invalid YAML
    """
    content = read_file(file_path, encoding)
    try:
        return yaml.safe_load(content)
    except yaml.YAMLError as e:
        logger.error(f"Error parsing YAML from {file_path}: {e}")
        raise


def read_csv_file(file_path: str, 
                 encoding: str = 'utf-8', 
                 delimiter: str = ',',
                 has_header: bool = True) -> List[Dict[str, str]]:
    """
    Read a CSV file and return its contents as a list of dictionaries.
    
    Args:
        file_path: Path to the CSV file
        encoding: File encoding
        delimiter: CSV delimiter
        has_header: Whether the CSV has a header row
        
    Returns:
        List of dictionaries representing the CSV rows
        
    Raises:
        FileNotFoundError: If the file does not exist
        csv.Error: If there's an error parsing the CSV
    """
    try:
        with open(file_path, 'r', encoding=encoding, newline='') as csvfile:
            if has_header:
                reader = csv.DictReader(csvfile, delimiter=delimiter)
                return list(reader)
            else:
                reader = csv.reader(csvfile, delimiter=delimiter)
                rows = list(reader)
                return [dict(zip([f"col{i}" for i in range(len(row))], row)) for row in rows]
    except csv.Error as e:
        logger.error(f"Error parsing CSV from {file_path}: {e}")
        raise


def read_excel_file(file_path: str, 
                   sheet_name: Optional[str] = None,
                   as_dict: bool = True) -> Union[List[Dict[str, Any]], pd.DataFrame]:
    """
    Read an Excel file and return its contents.
    
    Args:
        file_path: Path to the Excel file
        sheet_name: Name of the sheet to read (None for first sheet)
        as_dict: Whether to return the data as a list of dictionaries or a pandas DataFrame
        
    Returns:
        Excel data as a list of dictionaries or a pandas DataFrame
        
    Raises:
        FileNotFoundError: If the file does not exist
        ImportError: If pandas or openpyxl are not installed
        ValueError: If there's an error reading the Excel file
    """
    if not EXCEL_SUPPORT:
        raise ImportError("pandas and openpyxl are required for Excel support")
    
    try:
        if as_dict:
            data = pd.read_excel(file_path, sheet_name=sheet_name)
            return data.to_dict('records')
        else:
            return pd.read_excel(file_path, sheet_name=sheet_name)
    except Exception as e:
        logger.error(f"Error reading Excel file {file_path}: {e}")
        raise ValueError(f"Error reading Excel file: {e}")


def read_xml_file(file_path: str, encoding: str = 'utf-8') -> ET.Element:
    """
    Read an XML file and return its root element.
    
    Args:
        file_path: Path to the XML file
        encoding: File encoding
        
    Returns:
        Root element of the XML document
        
    Raises:
        FileNotFoundError: If the file does not exist
        ET.ParseError: If there's an error parsing the XML
    """
    try:
        tree = ET.parse(file_path)
        return tree.getroot()
    except ET.ParseError as e:
        logger.error(f"Error parsing XML from {file_path}: {e}")
        raise


# File Writing Operations

def write_file(file_path: str, 
               content: Union[str, bytes], 
               encoding: str = 'utf-8',
               binary: bool = False,
               create_dirs: bool = True) -> str:
    """
    Write content to a file.
    
    Args:
        file_path: Path to the file
        content: Content to write
        encoding: File encoding (for text content)
        binary: Whether to write in binary mode
        create_dirs: Whether to create parent directories if they don't exist
        
    Returns:
        The absolute path to the written file
        
    Raises:
        IOError: If there's an error writing to the file
    """
    if create_dirs:
        directory = os.path.dirname(os.path.abspath(file_path))
        ensure_directory_exists(directory)
    
    try:
        mode = 'wb' if binary else 'w'
        with open(file_path, mode=mode, encoding=None if binary else encoding) as file:
            file.write(content)
        logger.info(f"Successfully wrote to file: {file_path}")
        return os.path.abspath(file_path)
    except IOError as e:
        logger.error(f"Error writing to file {file_path}: {e}")
        raise


def write_json_file(file_path: str, 
                   data: Dict[str, Any], 
                   encoding: str = 'utf-8',
                   indent: int = 2,
                   create_dirs: bool = True) -> str:
    """
    Write data to a JSON file.
    
    Args:
        file_path: Path to the JSON file
        data: Data to write
        encoding: File encoding
        indent: JSON indentation level
        create_dirs: Whether to create parent directories if they don't exist
        
    Returns:
        The absolute path to the written file
        
    Raises:
        IOError: If there's an error writing to the file
        TypeError: If the data cannot be serialized to JSON
    """
    try:
        json_content = json.dumps(data, indent=indent, ensure_ascii=False)
        return write_file(file_path, json_content, encoding, create_dirs=create_dirs)
    except TypeError as e:
        logger.error(f"Error serializing data to JSON: {e}")
        raise


def write_yaml_file(file_path: str, 
                   data: Dict[str, Any], 
                   encoding: str = 'utf-8',
                   create_dirs: bool = True) -> str:
    """
    Write data to a YAML file.
    
    Args:
        file_path: Path to the YAML file
        data: Data to write
        encoding: File encoding
        create_dirs: Whether to create parent directories if they don't exist
        
    Returns:
        The absolute path to the written file
        
    Raises:
        IOError: If there's an error writing to the file
        yaml.YAMLError: If the data cannot be serialized to YAML
    """
    try:
        yaml_content = yaml.dump(data, default_flow_style=False, sort_keys=False)
        return write_file(file_path, yaml_content, encoding, create_dirs=create_dirs)
    except yaml.YAMLError as e:
        logger.error(f"Error serializing data to YAML: {e}")
        raise


def write_csv_file(file_path: str, 
                  data: List[Dict[str, Any]], 
                  encoding: str = 'utf-8',
                  delimiter: str = ',',
                  create_dirs: bool = True) -> str:
    """
    Write data to a CSV file.
    
    Args:
        file_path: Path to the CSV file
        data: List of dictionaries to write
        encoding: File encoding
        delimiter: CSV delimiter
        create_dirs: Whether to create parent directories if they don't exist
        
    Returns:
        The absolute path to the written file
        
    Raises:
        IOError: If there's an error writing to the file
        ValueError: If the data structure is invalid for CSV
    """
    if create_dirs:
        directory = os.path.dirname(os.path.abspath(file_path))
        ensure_directory_exists(directory)
    
    try:
        with open(file_path, 'w', encoding=encoding, newline='') as csvfile:
            if not data:
                # Create an empty file if data is empty
                writer = csv.writer(csvfile, delimiter=delimiter)
                writer.writerow([])
            else:
                fieldnames = data[0].keys()
                writer = csv.DictWriter(csvfile, fieldnames=fieldnames, delimiter=delimiter)
                writer.writeheader()
                writer.writerows(data)
        
        logger.info(f"Successfully wrote CSV to file: {file_path}")
        return os.path.abspath(file_path)
    except (IOError, ValueError) as e:
        logger.error(f"Error writing CSV to {file_path}: {e}")
        raise


def write_excel_file(file_path: str, 
                    data: Union[Dict[str, List[Dict[str, Any]]], List[Dict[str, Any]]], 
                    sheet_name: str = 'Sheet1',
                    create_dirs: bool = True) -> str:
    """
    Write data to an Excel file.
    
    Args:
        file_path: Path to the Excel file
        data: Data to write (either a dictionary mapping sheet names to data lists, 
              or a single list of dictionaries for one sheet)
        sheet_name: Default sheet name (used if data is a list)
        create_dirs: Whether to create parent directories if they don't exist
        
    Returns:
        The absolute path to the written file
        
    Raises:
        ImportError: If pandas or openpyxl are not installed
        IOError: If there's an error writing to the file
        ValueError: If the data structure is invalid for Excel
    """
    if not EXCEL_SUPPORT:
        raise ImportError("pandas and openpyxl are required for Excel support")
    
    if create_dirs:
        directory = os.path.dirname(os.path.abspath(file_path))
        ensure_directory_exists(directory)
    
    try:
        if isinstance(data, list):
            # Single sheet
            df = pd.DataFrame(data)
            df.to_excel(file_path, sheet_name=sheet_name, index=False)
        elif isinstance(data, dict):
            # Multiple sheets
            with pd.ExcelWriter(file_path) as writer:
                for sheet, sheet_data in data.items():
                    pd.DataFrame(sheet_data).to_excel(writer, sheet_name=sheet, index=False)
        else:
            raise ValueError("Data must be a list of dictionaries or a dictionary mapping sheet names to data lists")
        
        logger.info(f"Successfully wrote Excel to file: {file_path}")
        return os.path.abspath(file_path)
    except Exception as e:
        logger.error(f"Error writing Excel to {file_path}: {e}")
        raise ValueError(f"Error writing Excel file: {e}")


# File Operations

def copy_file(source_path: str, 
             destination_path: str, 
             create_dirs: bool = True) -> str:
    """
    Copy a file from source to destination.
    
    Args:
        source_path: Path to the source file
        destination_path: Path to the destination file
        create_dirs: Whether to create parent directories if they don't exist
        
    Returns:
        The absolute path to the destination file
        
    Raises:
        FileNotFoundError: If the source file does not exist
        IOError: If there's an error copying the file
    """
    if not os.path.exists(source_path):
        logger.error(f"Source file not found: {source_path}")
        raise FileNotFoundError(f"Source file not found: {source_path}")
    
    if create_dirs:
        directory = os.path.dirname(os.path.abspath(destination_path))
        ensure_directory_exists(directory)
    
    try:
        shutil.copy2(source_path, destination_path)
        logger.info(f"Successfully copied file from {source_path} to {destination_path}")
        return os.path.abspath(destination_path)
    except IOError as e:
        logger.error(f"Error copying file from {source_path} to {destination_path}: {e}")
        raise


def move_file(source_path: str, 
             destination_path: str, 
             create_dirs: bool = True) -> str:
    """
    Move a file from source to destination.
    
    Args:
        source_path: Path to the source file
        destination_path: Path to the destination file
        create_dirs: Whether to create parent directories if they don't exist
        
    Returns:
        The absolute path to the destination file
        
    Raises:
        FileNotFoundError: If the source file does not exist
        IOError: If there's an error moving the file
    """
    if not os.path.exists(source_path):
        logger.error(f"Source file not found: {source_path}")
        raise FileNotFoundError(f"Source file not found: {source_path}")
    
    if create_dirs:
        directory = os.path.dirname(os.path.abspath(destination_path))
        ensure_directory_exists(directory)
    
    try:
        shutil.move(source_path, destination_path)
        logger.info(f"Successfully moved file from {source_path} to {destination_path}")
        return os.path.abspath(destination_path)
    except IOError as e:
        logger.error(f"Error moving file from {source_path} to {destination_path}: {e}")
        raise


def delete_file(file_path: str, missing_ok: bool = False) -> bool:
    """
    Delete a file.
    
    Args:
        file_path: Path to the file
        missing_ok: Whether to ignore if the file does not exist
        
    Returns:
        True if the file was deleted, False if it did not exist and missing_ok is True
        
    Raises:
        FileNotFoundError: If the file does not exist and missing_ok is False
        IOError: If there's an error deleting the file
    """
    if not os.path.exists(file_path):
        if missing_ok:
            return False
        logger.error(f"File not found: {file_path}")
        raise FileNotFoundError(f"File not found: {file_path}")
    
    try:
        os.remove(file_path)
        logger.info(f"Successfully deleted file: {file_path}")
        return True
    except IOError as e:
        logger.error(f"Error deleting file {file_path}: {e}")
        raise


def create_temp_file(content: Union[str, bytes] = None, 
                    suffix: Optional[str] = None, 
                    prefix: Optional[str] = None, 
                    binary: bool = False) -> str:
    """
    Create a temporary file.
    
    Args:
        content: Optional content to write to the file
        suffix: Optional suffix for the file name
        prefix: Optional prefix for the file name
        binary: Whether the content is binary
        
    Returns:
        The path to the temporary file
        
    Raises:
        IOError: If there's an error creating the temporary file
    """
    try:
        with tempfile.NamedTemporaryFile(suffix=suffix, prefix=prefix, delete=False) as tmp:
            if content is not None:
                mode = 'wb' if binary else 'w'
                if not binary and isinstance(content, str):
                    content_bytes = content.encode('utf-8')
                else:
                    content_bytes = content
                tmp.write(content_bytes)
        
        logger.info(f"Successfully created temporary file: {tmp.name}")
        return tmp.name
    except IOError as e:
        logger.error(f"Error creating temporary file: {e}")
        raise


def create_temp_directory(suffix: Optional[str] = None, 
                         prefix: Optional[str] = None) -> str:
    """
    Create a temporary directory.
    
    Args:
        suffix: Optional suffix for the directory name
        prefix: Optional prefix for the directory name
        
    Returns:
        The path to the temporary directory
        
    Raises:
        IOError: If there's an error creating the temporary directory
    """
    try:
        temp_dir = tempfile.mkdtemp(suffix=suffix, prefix=prefix)
        logger.info(f"Successfully created temporary directory: {temp_dir}")
        return temp_dir
    except IOError as e:
        logger.error(f"Error creating temporary directory: {e}")
        raise


def get_file_hash(file_path: str, algorithm: str = 'sha256') -> str:
    """
    Calculate the hash of a file.
    
    Args:
        file_path: Path to the file
        algorithm: Hash algorithm to use ('md5', 'sha1', 'sha256', 'sha512')
        
    Returns:
        The hash of the file
        
    Raises:
        FileNotFoundError: If the file does not exist
        ValueError: If the algorithm is invalid
        IOError: If there's an error reading the file
    """
    import hashlib
    
    if not os.path.exists(file_path):
        logger.error(f"File not found: {file_path}")
        raise FileNotFoundError(f"File not found: {file_path}")
    
    hash_algorithms = {
        'md5': hashlib.md5,
        'sha1': hashlib.sha1,
        'sha256': hashlib.sha256,
        'sha512': hashlib.sha512
    }
    
    if algorithm not in hash_algorithms:
        valid_algorithms = ', '.join(hash_algorithms.keys())
        raise ValueError(f"Invalid hash algorithm: {algorithm}. Valid options are: {valid_algorithms}")
    
    hash_obj = hash_algorithms[algorithm]()
    
    try:
        with open(file_path, 'rb') as file:
            # Read the file in chunks to avoid loading large files into memory
            for chunk in iter(lambda: file.read(4096), b''):
                hash_obj.update(chunk)
        
        return hash_obj.hexdigest()
    except IOError as e:
        logger.error(f"Error reading file {file_path} for hashing: {e}")
        raise


def compare_files(file1_path: str, file2_path: str, binary: bool = False) -> bool:
    """
    Compare two files and check if they have the same content.
    
    Args:
        file1_path: Path to the first file
        file2_path: Path to the second file
        binary: Whether to compare the files in binary mode
        
    Returns:
        True if the files have the same content, False otherwise
        
    Raises:
        FileNotFoundError: If either file does not exist
        IOError: If there's an error reading the files
    """
    if not os.path.exists(file1_path):
        logger.error(f"First file not found: {file1_path}")
        raise FileNotFoundError(f"First file not found: {file1_path}")
    
    if not os.path.exists(file2_path):
        logger.error(f"Second file not found: {file2_path}")
        raise FileNotFoundError(f"Second file not found: {file2_path}")
    
    # Quick check: if the files have different sizes, they are not the same
    if os.path.getsize(file1_path) != os.path.getsize(file2_path):
        return False
    
    try:
        if binary:
            # Compare files byte by byte
            with open(file1_path, 'rb') as f1, open(file2_path, 'rb') as f2:
                for chunk1, chunk2 in zip(iter(lambda: f1.read(4096), b''), iter(lambda: f2.read(4096), b'')):
                    if chunk1 != chunk2:
                        return False
                # Check if one file has more data
                if f1.read(1) or f2.read(1):
                    return False
        else:
            # Compare text files line by line
            with open(file1_path, 'r', encoding='utf-8') as f1, open(file2_path, 'r', encoding='utf-8') as f2:
                for line1, line2 in zip(f1, f2):
                    if line1 != line2:
                        return False
                # Check if one file has more lines
                if next(f1, None) or next(f2, None):
                    return False
        
        return True
    except IOError as e:
        logger.error(f"Error comparing files {file1_path} and {file2_path}: {e}")
        raise


# Specialized file utilities for Watsonx IPG Testing project

def read_test_case_excel(file_path: str, sheet_name: Optional[str] = None) -> List[Dict[str, Any]]:
    """
    Read a test case Excel file in the standard format used by the project.
    
    Args:
        file_path: Path to the Excel file
        sheet_name: Name of the sheet to read (None for first sheet)
        
    Returns:
        List of test cases as dictionaries
        
    Raises:
        FileNotFoundError: If the file does not exist
        ImportError: If pandas or openpyxl are not installed
        ValueError: If the file doesn't follow the expected test case format
    """
    data = read_excel_file(file_path, sheet_name, as_dict=True)
    
    # Validate that the data follows the expected test case format
    required_fields = ['Test Case ID', 'Test Case Name', 'Description', 'Preconditions']
    for field in required_fields:
        if not all(field in item for item in data):
            raise ValueError(f"Invalid test case format: '{field}' field is missing")
    
    return data


def write_test_case_excel(file_path: str, 
                         test_cases: List[Dict[str, Any]], 
                         sheet_name: str = 'Test Cases',
                         template_path: Optional[str] = None) -> str:
    """
    Write test cases to an Excel file in the standard format used by the project.
    
    Args:
        file_path: Path to the output Excel file
        test_cases: List of test case dictionaries
        sheet_name: Name of the sheet to write
        template_path: Optional path to a template Excel file to use
        
    Returns:
        The absolute path to the written file
        
    Raises:
        ImportError: If pandas or openpyxl are not installed
        FileNotFoundError: If the template file does not exist
        ValueError: If the template doesn't follow the expected format
    """
    if not EXCEL_SUPPORT:
        raise ImportError("pandas and openpyxl are required for Excel support")
    
    if template_path:
        if not os.path.exists(template_path):
            raise FileNotFoundError(f"Template file not found: {template_path}")
        
        # Copy the template and modify it
        copy_file(template_path, file_path)
        
        # Load the workbook
        wb = openpyxl.load_workbook(file_path)
        
        # Check if the sheet exists, if not create it
        if sheet_name not in wb.sheetnames:
            wb.create_sheet(sheet_name)
        
        sheet = wb[sheet_name]
        
        # Clear existing data (keep header row)
        for row in range(2, sheet.max_row + 1):
            for col in range(1, sheet.max_column + 1):
                sheet.cell(row=row, column=col).value = None
        
        # Write the test cases
        df = pd.DataFrame(test_cases)
        for r_idx, row in enumerate(df.itertuples(index=False), 2):  # Start from the second row (after header)
            for c_idx, value in enumerate(row, 1):
                sheet.cell(row=r_idx, column=c_idx).value = value
        
        wb.save(file_path)
    else:
        # Create a new Excel file
        write_excel_file(file_path, test_cases, sheet_name)
    
    logger.info(f"Successfully wrote test cases to Excel file: {file_path}")
    return os.path.abspath(file_path)


def read_test_data_excel(file_path: str, sheet_name: Optional[str] = None) -> Dict[str, List[Dict[str, Any]]]:
    """
    Read a test data Excel file with multiple sheets for different test cases.
    
    Args:
        file_path: Path to the Excel file
        sheet_name: Name of specific data sheet to read (None for all sheets)
        
    Returns:
        Dictionary mapping test case IDs to lists of test data rows
        
    Raises:
        FileNotFoundError: If the file does not exist
        ImportError: If pandas or openpyxl are not installed
    """
    if not EXCEL_SUPPORT:
        raise ImportError("pandas and openpyxl are required for Excel support")
    
    xlsx = pd.ExcelFile(file_path)
    
    if sheet_name is not None:
        # Read a specific sheet
        if sheet_name not in xlsx.sheet_names:
            raise ValueError(f"Sheet '{sheet_name}' not found in file {file_path}")
        
        data = pd.read_excel(xlsx, sheet_name=sheet_name)
        return {sheet_name: data.to_dict('records')}
    else:
        # Read all sheets
        result = {}
        for sheet in xlsx.sheet_names:
            data = pd.read_excel(xlsx, sheet_name=sheet)
            result[sheet] = data.to_dict('records')
        
        return result


def save_execution_report(report_data: Dict[str, Any], 
                         output_dir: str,
                         report_name: Optional[str] = None,
                         timestamp_format: str = '%Y%m%d_%H%M%S') -> str:
    """
    Save an execution report to a file.
    
    Args:
        report_data: Report data
        output_dir: Directory to save the report
        report_name: Optional report name (will be auto-generated if None)
        timestamp_format: Format for the timestamp in the filename
        
    Returns:
        The path to the saved report
        
    Raises:
        IOError: If there's an error writing the report
    """
    ensure_directory_exists(output_dir)
    
    timestamp = datetime.now().strftime(timestamp_format)
    
    if report_name is None:
        # Generate a name based on the report content if possible
        test_case_id = report_data.get('test_case_id', '')
        execution_id = report_data.get('execution_id', '')
        status = report_data.get('status', '')
        
        if test_case_id and execution_id:
            report_name = f"report_{test_case_id}_{execution_id}_{status}_{timestamp}.json"
        else:
            report_name = f"execution_report_{timestamp}.json"
    
    report_path = os.path.join(output_dir, report_name)
    
    # Add timestamp to the report data
    report_data['generated_at'] = datetime.now().isoformat()
    
    return write_json_file(report_path, report_data)


def load_jira_export(file_path: str) -> List[Dict[str, Any]]:
    """
    Load and parse a JIRA export file (CSV or Excel).
    
    Args:
        file_path: Path to the JIRA export file
        
    Returns:
        List of JIRA issues as dictionaries
        
    Raises:
        FileNotFoundError: If the file does not exist
        ValueError: If the file format is not supported
    """
    ext = get_file_extension(file_path)
    
    if ext == 'csv':
        return read_csv_file(file_path)
    elif ext in ['xlsx', 'xls']:
        return read_excel_file(file_path)
    else:
        raise ValueError(f"Unsupported JIRA export format: {ext}. Expected: csv, xlsx, or xls")


def extract_requirements_from_doc(file_path: str) -> List[Dict[str, str]]:
    """
    Extract requirements from a document file.
    Simple implementation - in a real system, this would use NLP or other techniques.
    
    Args:
        file_path: Path to the document file
        
    Returns:
        List of extracted requirements as dictionaries
        
    Raises:
        FileNotFoundError: If the file does not exist
        ValueError: If the file format is not supported
    """
    ext = get_file_extension(file_path)
    
    # This is a simplified implementation - in reality, this would involve
    # more sophisticated document parsing, possibly using NLP techniques
    
    if ext == 'txt':
        # Simple text file - assume each line is a requirement
        content = read_file(file_path)
        lines = [line.strip() for line in content.split('\n') if line.strip()]
        
        return [
            {
                'id': f'REQ_{i+1:03d}',
                'description': line,
                'source': file_path,
                'line_number': i+1
            }
            for i, line in enumerate(lines)
        ]
    
    elif ext in ['docx', 'doc']:
        # Would typically use a library like python-docx
        # For simplicity, we'll just return a placeholder message
        return [{
            'id': 'DOCX_PLACEHOLDER',
            'description': 'Document parsing not implemented in this example',
            'source': file_path
        }]
    
    elif ext == 'pdf':
        # Would typically use a library like PyPDF2 or pdfminer
        # For simplicity, we'll just return a placeholder message
        return [{
            'id': 'PDF_PLACEHOLDER',
            'description': 'PDF parsing not implemented in this example',
            'source': file_path
        }]
    
    else:
        raise ValueError(f"Unsupported document format: {ext}. Expected: txt, docx, doc, or pdf")


def safe_filename(name: str) -> str:
    """
    Convert a string to a safe filename.
    
    Args:
        name: The string to convert
        
    Returns:
        A safe filename
    """
    # Replace spaces with underscores
    s = name.replace(' ', '_')
    
    # Remove invalid characters
    s = ''.join(c for c in s if c.isalnum() or c in '._-')
    
    # Limit length
    s = s[:255]
    
    return s


def get_mime_type(file_path: str) -> str:
    """
    Get the MIME type of a file.
    
    Args:
        file_path: Path to the file
        
    Returns:
        The MIME type
        
    Raises:
        FileNotFoundError: If the file does not exist
    """
    import mimetypes
    
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"File not found: {file_path}")
    
    mime_type, _ = mimetypes.guess_type(file_path)
    return mime_type or 'application/octet-stream'


def create_backup(file_path: str, backup_dir: Optional[str] = None) -> str:
    """
    Create a backup of a file.
    
    Args:
        file_path: Path to the file
        backup_dir: Directory to store backups (uses file's directory if None)
        
    Returns:
        The path to the backup file
        
    Raises:
        FileNotFoundError: If the file does not exist
        IOError: If there's an error creating the backup
    """
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"File not found: {file_path}")
    
    if backup_dir is None:
        backup_dir = os.path.dirname(os.path.abspath(file_path))
    else:
        ensure_directory_exists(backup_dir)
    
    # Get the file name and create a backup name with timestamp
    file_name = os.path.basename(file_path)
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    backup_name = f"{os.path.splitext(file_name)[0]}_{timestamp}{os.path.splitext(file_name)[1]}"
    backup_path = os.path.join(backup_dir, backup_name)
    
    return copy_file(file_path, backup_path)


def detect_encoding(file_path: str) -> str:
    """
    Detect the encoding of a text file.
    
    Args:
        file_path: Path to the file
        
    Returns:
        The detected encoding
        
    Raises:
        FileNotFoundError: If the file does not exist
        IOError: If there's an error reading the file
    """
    try:
        import chardet
        
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"File not found: {file_path}")
        
        # Read a sample of the file to detect encoding
        with open(file_path, 'rb') as f:
            sample = f.read(4096)
            if len(sample) == 0:
                return 'utf-8'  # Default to utf-8 for empty files
            
        result = chardet.detect(sample)
        encoding = result['encoding']
        confidence = result['confidence']
        
        logger.info(f"Detected encoding for {file_path}: {encoding} (confidence: {confidence:.2f})")
        
        return encoding or 'utf-8'
    except ImportError:
        logger.warning("chardet library not installed. Defaulting to UTF-8 encoding")
        return 'utf-8'
    except IOError as e:
        logger.error(f"Error reading file {file_path} for encoding detection: {e}")
        raise


def merge_csv_files(file_paths: List[str], 
                   output_path: str,
                   encoding: str = 'utf-8',
                   delimiter: str = ',') -> str:
    """
    Merge multiple CSV files into one.
    
    Args:
        file_paths: List of paths to CSV files
        output_path: Path to the output CSV file
        encoding: File encoding
        delimiter: CSV delimiter
        
    Returns:
        The path to the merged CSV file
        
    Raises:
        FileNotFoundError: If any of the input files does not exist
        ValueError: If the headers of the CSV files don't match
        IOError: If there's an error reading or writing the files
    """
    if not file_paths:
        raise ValueError("No input files provided")
    
    # Check that all input files exist
    for file_path in file_paths:
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"File not found: {file_path}")
    
    # Read the first file to get the headers
    with open(file_paths[0], 'r', encoding=encoding, newline='') as f:
        reader = csv.reader(f, delimiter=delimiter)
        headers = next(reader)
    
    # Create the output file and write the headers
    with open(output_path, 'w', encoding=encoding, newline='') as out_file:
        writer = csv.writer(out_file, delimiter=delimiter)
        writer.writerow(headers)
        
        # Process each input file
        for file_path in file_paths:
            with open(file_path, 'r', encoding=encoding, newline='') as in_file:
                reader = csv.reader(in_file, delimiter=delimiter)
                
                # Verify the headers match
                file_headers = next(reader)
                if file_headers != headers:
                    raise ValueError(f"Headers in {file_path} don't match the headers in {file_paths[0]}")
                
                # Copy the rows
                for row in reader:
                    writer.writerow(row)
    
    logger.info(f"Successfully merged {len(file_paths)} CSV files into {output_path}")
    return os.path.abspath(output_path)


def search_text_in_files(directory_path: str, 
                        search_text: str,
                        extensions: Optional[List[str]] = None,
                        recursive: bool = True,
                        case_sensitive: bool = False) -> List[Dict[str, Any]]:
    """
    Search for text in files.
    
    Args:
        directory_path: Path to the directory to search in
        search_text: Text to search for
        extensions: List of file extensions to search in (None for all)
        recursive: Whether to search recursively in subdirectories
        case_sensitive: Whether the search is case-sensitive
        
    Returns:
        List of matches with file path and line number
        
    Raises:
        FileNotFoundError: If the directory does not exist
        IOError: If there's an error reading a file
    """
    if not os.path.exists(directory_path):
        raise FileNotFoundError(f"Directory not found: {directory_path}")
    
    matches = []
    
    # Get all files to search
    file_paths = list_files(directory_path, extensions, recursive)
    
    for file_path in file_paths:
        try:
            # Try to detect the encoding
            try:
                encoding = detect_encoding(file_path)
            except:
                # If encoding detection fails, default to utf-8
                encoding = 'utf-8'
            
            with open(file_path, 'r', encoding=encoding, errors='replace') as f:
                for i, line in enumerate(f, 1):
                    if case_sensitive:
                        if search_text in line:
                            matches.append({
                                'file_path': file_path,
                                'line_number': i,
                                'line': line.rstrip()
                            })
                    else:
                        if search_text.lower() in line.lower():
                            matches.append({
                                'file_path': file_path,
                                'line_number': i,
                                'line': line.rstrip()
                            })
        except UnicodeDecodeError:
            # Skip binary files
            logger.debug(f"Skipping binary file: {file_path}")
            continue
        except IOError as e:
            logger.error(f"Error reading file {file_path}: {e}")
            raise
    
    return matches